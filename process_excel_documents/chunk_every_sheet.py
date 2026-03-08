import json
import os
import pandas as pd
from transformers import AutoTokenizer
import sys
from pathlib import Path
# Ajouter le répertoire parent au path
parent_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

from my_paths import BGE_M3_TOKENIZER_PATH

def read_excel_sheet(filepath, sheetname):
    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.xls':
        with open(filepath, 'rb') as f:
            df = pd.read_excel(f, sheet_name=sheetname, engine='xlrd', header=None)
    elif ext == '.xlsx':
        with open(filepath, 'rb') as f:
            df = pd.read_excel(f, sheet_name=sheetname, engine='openpyxl', header=None)
    else:
        raise ValueError(f"Format de fichier non supporté : {filepath}")

    # Retirer colonnes/rows entièrement vides
    df = df.dropna(axis=1, how='all').dropna(how='all')

    # Renommer colonnes : si NaN → A, B, C...
    df.columns = [
        col if not (isinstance(col, str) and col.startswith("Unnamed"))
        else chr(65 + i)
        for i, col in enumerate(df.columns)
    ]

    # Restaurer les numéros de lignes Excel authentiques
    df.index = df.index + 1

    return df

def split_df_by_token_limit(df, tokenizer, header_lines, max_tokens, doc_name, sheetname, max_tokens_embedding_model):
    chunks = []
    n = len(df)
    print(f"Header lines: {header_lines}")
    header_indices = list(range(1, header_lines+1)) #+1 dans le range car on a un index qui commence à 1
    header_df = df.loc[header_indices]      # Vu qu'on a supprimé les lignes vides du header on utilise loc pour sélectionner avec l'index et pas la position
    print(f"Header df : {header_df}")
    start_idx = header_lines                # Première ligne de données

    while start_idx < n:
        end_idx = start_idx
        last_valid_end = start_idx

        while end_idx < n:
            chunk_df = pd.concat([header_df, df.iloc[start_idx:end_idx + 1]])
            csv_str = chunk_df.to_csv(index=True)

            n_tokens = len(tokenizer(csv_str)['input_ids'])
            if n_tokens <= max_tokens:
                last_valid_end = end_idx
                end_idx += 1
            elif last_valid_end == start_idx and n_tokens > max_tokens_embedding_model:
                print(f"Trop de tokens : {n_tokens} > {max_tokens_embedding_model} pour la ligne {start_idx} de la feuille {sheetname} du document {doc_name}. Impossible de chunker. Skip.")
                break
            else :
                break 

        if not(last_valid_end == start_idx and n_tokens > max_tokens_embedding_model) :
            # Chunk final
            chunk_df = pd.concat([header_df, df.iloc[start_idx:last_valid_end + 1]])

            chunks.append(chunk_df)

        start_idx = last_valid_end + 1

    return chunks

import json
import os
import pandas as pd
import xlrd
import openpyxl
from transformers import AutoTokenizer
from tqdm import tqdm
import xlwings as xw

def get_freeze_panes_with_filter(file_path, sheet_name):

    app = xw.App(visible=False)
    try:
        wb = app.books.open(file_path, read_only=True)
        ws = wb.sheets[sheet_name]
        ws.activate()
        wb.api.Windows(1).Activate()
        aw = ws.api.Application.ActiveWindow

        split_row = aw.SplitRow if aw.FreezePanes else 0

        # Vérification presenza filtre auto
        af = ws.api.AutoFilter
        if af is not None:
            filter_row = af.Range.Row  # ligne du filtre (souvent dernière ligne header)
            # Si la ligne filtre > split_row, on adapte
            if filter_row > split_row:
                split_row = filter_row

        wb.close()
        return int(split_row)

    finally:
        app.quit()

def sheet_already_processed(data, doc_name, sheetname, rel_filepath) :
    for entry in data :
        if entry['metadata']['sheet_name'] == sheetname and entry['metadata']['doc_name'] == doc_name and entry['metadata']['doc_path'] == rel_filepath:
            return(True)
    return(False)

from mistral_langchain_wrapper import MistralChatWrapper
from dotenv import load_dotenv

load_dotenv()
API_KEY = os.getenv('API_KEY')
llm = MistralChatWrapper(api_key=API_KEY, model="medium")

def correct_frozen_lines(file_name, sheet_name, file_path) :
    print(f"\n---\nTraitement pour '{file_name}' feuille '{sheet_name}' : nombre de lignes figées non renseigné ou = 0. \nOn demande à un LLM de trouver les potentielles lignes d'en-tête.")

    # Lire les 15 premières lignes
    try:
        preview_df = read_excel_sheet(file_path, sheetname=sheet_name)
    except Exception as e:
        print(f"Erreur lecture Excel {file_path} / feuille {sheet_name} : {e}")
    print(f'Première lignes : {preview_df.head(15)}')
    prompt = f"""Your task is to determine the number of header rows in a given Excel file, based on the content of the first few rows.  
Note that since empty rows have been removed, the row indices are not continuous.
The header rows always begin at the first row of the file.
You must return the INDEX of the last header row.
If there are no clear header rows, you MUST output 0.
Output ONLY the index of the last header row—no explanations, just the number.

The first rows of the Excel file :
{preview_df.head(15)}
"""
    n_frozen_corrected = int(llm.invoke(prompt).content)
    print(f'Réponse du LLM : {n_frozen_corrected}')
    return(n_frozen_corrected)

def get_headers(headers_output_json_path, raw_data_dir) :
    erreurs = 0
    feuilles_vides = 0
    if not os.path.exists(headers_output_json_path):
        headers = {}
    else : 
        with open(headers_output_json_path, 'r', encoding='utf-8') as f:
            headers = json.load(f)
    # Charge les statistiques de dossiers
    for root, dirs, files in tqdm(os.walk(raw_data_dir)):
        for file in files:
            if file.lower().endswith(('.xlsx', '.xls')):
                filepath = os.path.join(root, file)
                rel_filepath = str(Path(filepath).relative_to(raw_data_dir))
                doc_name = os.path.basename(filepath)
                # Lire le fichier Excel
                if file.lower().endswith('.xlsx'):
                    print(f"Ouverture du fichier {filepath}")
                    wb = openpyxl.load_workbook(filepath, read_only=True)
                    sheetnames = wb.sheetnames
                else:
                    print(f"Ouverture du fichier {filepath}")
                    wb = xlrd.open_workbook(filepath, on_demand=True)
                    sheetnames = wb.sheet_names()
                # Analyser les feuilles
                for sheetname in sheetnames:
                    # Vérifier si la feuille a déjà été traitée
                    if rel_filepath in headers : 
                        if sheetname in headers[rel_filepath] :
                            print(f"Headers de la feuille déjà repérés : {filepath} :: {sheetname}, saut.")
                            continue

                    try:
                        # Lire la feuille Excel en DataFrame
                        df = read_excel_sheet(filepath, sheetname)

                        # Si la feuille est vide, l'ignorer
        
                        if df.empty:
                            feuilles_vides += 1
                            print(f"  Feuille vide après nettoyage, ignorée.")
                            continue

                        # Récupérer les lignes figées
                        try : 
                            split_row = get_freeze_panes_with_filter(filepath, sheetname)
                        except Exception as e : 
                            print(f"Problème lors de la tentative d'obtention des lignes figées : {e}. Valeur par défaut mise. (0)")
                            split_row = 0
                        print(f"  Lignes figées : {split_row}")
                        if split_row == 0 :
                            split_row = correct_frozen_lines(doc_name, sheetname, filepath)
                        if rel_filepath in headers :
                            headers[rel_filepath][sheetname] = split_row
                        else :
                            headers[rel_filepath] = {}
                            headers[rel_filepath][sheetname] = split_row
                        with open(headers_output_json_path, 'w', encoding='utf-8') as f_out:
                            json.dump(headers, f_out, ensure_ascii=False, indent=2)

                    except Exception as e:
                        erreurs += 1
                        print(f"Erreur sur {doc_name}::{sheetname} : {e}")

    # Sauvegarder les résultats dans le fichier JSON
    with open(headers_output_json_path, 'w', encoding='utf-8') as f_out:
        json.dump(headers, f_out, ensure_ascii=False, indent=2)
    
    return(headers)

def export_sheets_chunked(output_json_path, headers_output_json_path, raw_data_dir, max_tokens=8000, max_sub_chunks_tokens=2000):
    """
    Exporte les feuilles sous forme de chunks avec les critères donnés (max 8192 tokens).
    Le fichier `output_json_path` contient les chunks à exporter.
    """
    headers = get_headers(headers_output_json_path, raw_data_dir)
    if not os.path.exists(output_json_path):
        chunks = []
    else : 
        with open(output_json_path, 'r', encoding='utf-8') as f:
            chunks = json.load(f)
    erreurs = 0
    feuilles_vides = 0
    tokenizer = AutoTokenizer.from_pretrained(BGE_M3_TOKENIZER_PATH)

    # Charge les statistiques de dossiers
    for root, dirs, files in tqdm(os.walk(raw_data_dir)):
        for file in files:
            if file.lower().endswith(('.xlsx', '.xls')):
                filepath = os.path.join(root, file)
                rel_filepath = str(Path(filepath).relative_to(raw_data_dir))
                doc_name = os.path.basename(filepath)
                # Lire le fichier Excel
                if file.lower().endswith('.xlsx'):
                    wb = openpyxl.load_workbook(filepath, read_only=True)
                    sheetnames = wb.sheetnames
                else :
                    wb = xlrd.open_workbook(filepath, on_demand=True)
                    sheetnames = wb.sheet_names()
                # Analyser les feuilles
                for sheetname in sheetnames:
                    # Vérifier si la feuille a déjà été traitée
                    if sheet_already_processed(chunks, doc_name, sheetname, rel_filepath):
                        print(f"Feuille déjà traitée : {filepath} :: {sheetname}, saut.")
                        continue

                    try:
                        print(f"Traitement de {doc_name} :: {sheetname}")

                        # Lire la feuille Excel en DataFrame
                        df = read_excel_sheet(filepath, sheetname)

                        # Si la feuille est vide, l'ignorer
                        if df.empty:
                            feuilles_vides += 1
                            print(f"  Feuille vide après nettoyage, ignorée.")
                            continue

                        # Convertir le DataFrame en CSV pour vérifier la taille en tokens
                        full_csv = df.to_csv(index=True)
                        provenance = f"From document: {doc_name}, sheet: {sheetname}"
                        chunk_content = provenance + "\n" + full_csv
                        total_tokens = len(tokenizer(chunk_content)['input_ids'])

                        # Si le total_tokens ne dépasse pas la limite
                        if total_tokens <= max_tokens:
                            chunk = {
                                "content": chunk_content,
                                "metadata": {
                                    "doc_name": doc_name,
                                    "doc_path": rel_filepath,
                                    "sheet_name": sheetname,
                                    "chunk_index": 0,
                                    "total_chunks": 1,
                                    "token_count": total_tokens
                                }
                            }
                            chunks.append(chunk)
                            # Sauvegarder les résultats dans le fichier JSON
                            with open(output_json_path, 'w', encoding='utf-8') as f_out:
                                json.dump(chunks, f_out, ensure_ascii=False, indent=2)
                            
                        else:
                             # Récupérer les lignes figées
                            split_row = headers[rel_filepath][sheetname]
                            # Découper la feuille en plusieurs chunks
                            df_chunks = split_df_by_token_limit(df, tokenizer, header_lines=split_row, max_tokens=max_sub_chunks_tokens, doc_name=doc_name, sheetname=sheetname, max_tokens_embedding_model=max_tokens)
                            total_chunks = len(df_chunks)

                            for idx, chunk_df in enumerate(df_chunks):
                                chunk_csv = chunk_df.to_csv(index=True)
                                provenance = f"From document: {doc_name}, sheet: {sheetname}"
                                chunk_content = provenance + "\n" + chunk_csv
                                token_count = len(tokenizer(chunk_content)['input_ids'])

                                chunk = {
                                    "content": chunk_content,
                                    "metadata": {
                                        "doc_name": doc_name,
                                        "doc_path": rel_filepath,
                                        "sheet_name": sheetname,
                                        "chunk_index": idx,
                                        "total_chunks": total_chunks,
                                        "token_count": token_count
                                    }
                                }
                                chunks.append(chunk)
                            
                            # Sauvegarder les résultats dans le fichier JSON, et cela que si on a traité toute la feuille
                            with open(output_json_path, 'w', encoding='utf-8') as f_out:
                                json.dump(chunks, f_out, ensure_ascii=False, indent=2)

                    except Exception as e:
                        erreurs += 1
                        print(f"Erreur sur {doc_name}::{sheetname} : {e}")

    # Sauvegarder les résultats dans le fichier JSON
    with open(output_json_path, 'w', encoding='utf-8') as f_out:
        json.dump(chunks, f_out, ensure_ascii=False, indent=2)

    print(f"\nTraitement terminé. {len(chunks)} chunks exportés.")
    print(f"{feuilles_vides} feuilles vides ignorées, {erreurs} erreurs rencontrées.")

if __name__ == "__main__":
    import sys

    # Ajouter le répertoire parent au path
    parent_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    if parent_dir not in sys.path:
        sys.path.insert(0, parent_dir)

    from my_paths import BASE_PATH, RAW_DATA_DIR, PARSED_DATA_DIR
    json_stats = BASE_PATH / "process_excel_documents/stats.json"
    header_file_path = PARSED_DATA_DIR / "headers.json"
    get_headers(header_file_path, RAW_DATA_DIR)
